import io
import datetime
from openpyxl import Workbook
from tortoise.fields.relational import BackwardFKRelation
from core.database.models import Report, Exhibit


async def create_excel(model):
    entries = await model.all()

    file_in_memory = io.BytesIO()
    book = Workbook()
    sheet = book.active

    # get model headers
    headers = []
    for field in model._meta.fields_map.values():
        if type(field) != BackwardFKRelation:
            headers.append(field.model_field_name)
    sheet.append(headers)

    # add users data
    for entry in entries:
        row = []
        for field_name in headers:
            cell = getattr(entry, field_name)
            if type(cell) == datetime.datetime:
                cell: datetime.datetime = cell.replace(tzinfo=None)
            row.append(cell)

        sheet.append(row)

    book.save(file_in_memory)
    file_in_memory.seek(0)

    return file_in_memory


async def create_excel_after_checking(reports: list[Report]):
    file_in_memory = io.BytesIO()
    book = Workbook()
    sheet = book.active

    sheet.append(['Айди отчета', 'Экспонат', 'Статус', 'Комментарий', 'ФИО проверяющего'])

    # add data
    is_empty = True
    for report in reports:
        if report.status != Report.StatusType.work:
            is_empty = False
            sheet.append(
                [
                    report.id,
                    (await report.exhibit).name,
                    report.status.value,
                    report.description,
                    (await report.creator).fio
                ]
            )

    book.save(file_in_memory)
    file_in_memory.seek(0)

    return file_in_memory, is_empty


async def sort_reports_by_date():
    reports = await Report.all().order_by('created_at')  # Получаем все записи и сортируем по дате создания

    report_data = {}
    for report in reports:
        key = (report.created_at.date(), (await report.exhibit).name)
        report_data[key] = report.status.value

    for date in report_data.items():
        print(date)

    data_dict = report_data
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active

    # Записываем даты в первую строку
    dates = sorted(set(date for date, _ in data_dict.keys()))
    ws.append([''] + [str(date) for date in dates])

    # Записываем статусы по каждой комбинации (дата, название)
    exhibit_names = sorted(set(name for _, name in data_dict.keys()))
    for name in exhibit_names:
        row = [name]
        for date in dates:
            status = data_dict.get((date, name), '')
            row.append(status)
        ws.append(row)

    # Сохраняем книгу в файл
    wb.save('exhibit_statuses.xlsx')
